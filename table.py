import pandas as pd
import win32com.client as win32

from comfunc import *
from tkinter.constants import END

from comfunc import search_file
from openpyxl import load_workbook
from openpyxl.worksheet.table import *
from openpyxl.utils import *

def DataType_set (DataType):
    if DataType[0] == '#':
        res = 'object'        
    elif DataType.find ('int') > 0 or DataType.find ('ref_id') > 0 :
        res = 'int64'
    elif DataType.find('float') > 0:
        res = 'float'
    else :
        res = 'object'
    return res


class excel ():
    def __init__ (self, wb):
        self.wb = wb
        self.wbpath = search_file(wb)
    def openedxl (self):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        return excel
    def dispatchxl (self):
        excel = win32.Dispatch("Excel.Application")
        if excel.Visible == False:
            excel.Visible = True
        return excel
    def tablelistxl (self):
        tbls = list()
        wb = load_workbook (self.wbpath)
        for sht in wb.sheetnames:
            for tbl in wb[sht]._tables:
                if tbl.find('____') == -1 :
                    tbls.append (tbl)
        return tbls
    # def readtbl (self):
    def readxltbl(self, tblname):
        wb = load_workbook (self.wbpath)
        for sht in wb.sheetnames:
            for table in wb[sht]._tables:
                if tblname == table:
                    print ('table : {}'.format(table))
                    print ('sht: {}'.format(sht))
                    ws = wb[sht]
                    break
        table = ws.tables[tblname]
        table_range = table.ref

        table_head = ws[table_range][0]
        table_data = ws[table_range][1:]

        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}
        for row in table_data:
            row_val = [cell.value for cell in row]
            for key, val in zip(columns, row_val):
                data[key].append(val)
        return columns, data

    def readxltbl_df (self, tblname):
        columns, data = readxltbl(tblname)
        df = pd.DataFrame(data=data, columns=columns, dtype='object')        
        return df

    def savexl (self): # 이건 쓸 수 있나?
        excel = self.openedxl()
        saved = None
        for active_wb in excel.Workbooks:
            if self.wb == active_wb.Name:
                saved = active_wb.Name
                active_wb.Save()
        return saved
    def openxl(self):
        excel = self.openedxl()
        for active_wb in excel.Workbooks:
            if self.wb == active_wb.Name:
                return -1
        excel = self.dispatchxl()
        excel.Workbooks.Open(self.wbpath)
    def instancexl (self): # 이건 쓸 수 있나?
        excel = self.dispatchxl()
        excel.Workbooks.Add(self.wbpath)

class sheettable ():
    def __init__ (self, wb, ws, tbl):
        filepath = search_file (wb)
        self.wb = load_workbook (filepath)
        self.ws = self.wb[ws]
        self.tbl = self.ws.tables[tbl]

        self.dtbl_df = self.read_desc_tbl()
        self.mtbl_columns, self.mtbl_data, self.mtbl_df = self.read_main_tbl()

    def read_desc_tbl (self): # 행열 전환
        tbl_range = self.tbl.ref

        sCell = tbl_range.split(':')[0]
        eCell = tbl_range.split(':')[1]

        tbl_sCol = sCell[:get_numeric_pos (sCell)]
        tbl_eCol = eCell[:get_numeric_pos (eCell)]
        tbl_sRow = int(sCell[get_numeric_pos (sCell):])
        tbl_range = str ( tbl_sCol + '1:' + tbl_eCol + str(tbl_sRow))

        #일반
        table_head = self.ws[tbl_range][tbl_sRow-1]
        table_data = self.ws[tbl_range][0:tbl_sRow-2]
        
        columns, data = self.read_tbl (table_head, table_data)
        print (columns)
        tbl_df = pd.DataFrame(data=data, columns=columns)
        
        return tbl_df.set_index('Comment').T

    def read_main_tbl(self):
        tbl_range = self.tbl.ref
        table_head = self.ws[tbl_range][0]
        table_data = self.ws[tbl_range][1:]
        columns, data = self.read_tbl (table_head, table_data)

        tbl_df = pd.DataFrame(data=data, columns=columns, dtype='object')

        # Comment 소거 및, 주석 Row 소거
        skiprows =[]
        for i, x in enumerate (tbl_df.iloc[:,0]):
            if x is not None :
                if x.find ('//') == 0 :
                    print ('i : {}'.format(i))
                    print ('x : {}'.format(x))
                    print ('tbl_df.index[{}] :'.format(tbl_df.index[i]))
                    skiprows.append(i)
                    # tbl_df.drop(tbl_df.index[i], inplace = True)
        [tbl_df.drop(tbl_df.index[i], inplace = True) for i in reversed (skiprows)] # revers 때문에 이렇게 처리. 더 좋은 방법 알아보자
                            
        # 코멘트 소거
        tbl_df.drop (['Comment'], axis = 'columns', inplace = True)

        # 데이터 타입 변경 / 안해도 될
        for i, x in enumerate (self.dtbl_df['DataType']):
            col = self.dtbl_df.index[i]
            res = DataType_set (x)
            print ('target : {}'.format(res))

            # tbl_df = tbl_df.astype({col : res})


        # C&S != n 케이스 소거
        for i, x in enumerate (self.dtbl_df['C&S']):
            col = self.dtbl_df.index[i]
            # print ('id : {}'.format(i))
            # print ('col : {}'.format(col)) #a가 아닌 컬럼 명
            # print ('val : {}'.format(x))
            if x != 'a':
                print ('소거 : {}'.format(col))
                tbl_df.drop([col], axis = 'columns', inplace = True)

        # ID 설정
        key = self.dtbl_df.index[0]
        tbl_df.set_index(key, inplace = True) # id 가 아닌 경우 체크

        return columns, data, tbl_df

    def read_tbl (self, table_head, table_data):
        columns = [column.value for column in table_head]
        data = {column: [] for column in columns}
        for row in table_data:
            row_val = [cell.value for cell in row]
            for key, val in zip(columns, row_val):
                data[key].append(val)

        return columns, data

    def export (self, path):
        res = self.mtbl_df.to_csv (path + self.ws.title + '.csv')
        print (res)

    def validate (self):
        validate_lst = ['DataType',
                'Ref0',
                'Ref1',
                'MaxArrayCount',
                'MinValue',
                'MaxValue',
                'DescPath']

        for x in validate_lst:
            print (self.dtbl_df[x])
            for i, x in enumerate (self.dtbl_df['C&S']):
                col = self.dtbl_df.index[i]
                print ('id : {}'.format(i))
                print ('col : {}'.format(col)) #a가 아닌 컬럼 명
                print ('val : {}'.format(x))

class csv_table ():
    pass


